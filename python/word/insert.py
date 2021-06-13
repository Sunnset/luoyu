from docxtpl import DocxTemplate
from openpyxl import load_workbook
import os
def replace(obj):
    if obj is None:
       obj = ''
       return obj
 # 加载要填入的数据
wb = load_workbook(r"source.xlsx")
ws = wb['Sheet1']
contexts = []
for row in range(1, ws.max_row + 1):
    name = ws["A" + str(row)].value
    
    context = {"name": name}
    contexts.append(context)
contexts
# 创建要保存的文件夹
os.mkdir("word")
for context in contexts:
    print(context)
    tpl = DocxTemplate(r"template.docx")
    tpl.render(context)
    tpl.save("{}-.docx".format(context["name"]))